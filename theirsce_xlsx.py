#!/usr/bin/env python3
# ============================================================
#  theirsce_xlsx.py
#
#  DAT의 모든(또는 지정) THEIRSCE 씬을 읽어 엑셀로 정리 / 엑셀->번역JSON 역주입.
#  PS2 매핑 엑셀 스타일: 씬마다 헤더행 + 그 아래 대사행.
#
#  export: JP DAT -> 씬별 일본어 추출 -> 한 엑셀에 정리 (Korean 열 비움).
#          기존 translation/{씬}.json 의 kr이 있으면 미리 채워 진행상황 표시.
#  import: 채운 엑셀 -> translation/{씬}.json 일괄 생성/갱신 (kr 주입).
#          build_dat --all 로 바로 빌드 가능.
#
#  in-process 추출(MiniTOR 1회 생성 후 재사용)이라 826씬도 빠름.
#
#  사용:
#    py theirsce_xlsx.py export --out tor_kr.xlsx --all
#    py theirsce_xlsx.py export --out tor_kr.xlsx --scenes 4234,4246,4247
#    py theirsce_xlsx.py import --in tor_kr.xlsx
# ============================================================
import argparse
import json
import os
import struct
import sys
from pathlib import Path

PYTHONLIB_PATH = r"D:\PythonLib"
PTR = 0x126F90
STORY_OFFSET = 6001  # PS2 Story번호 = 씬번호 + 6001

COLS = ["Scene", "Id", "Speaker", "Voice", "Type", "Japanese", "Korean"]


def read_pointers(eboot_path, dat_size, ptr_base=PTR):
    eboot = Path(eboot_path).read_bytes()
    ptrs, i = [], 0
    while True:
        off = ptr_base + i * 4
        if off + 4 > len(eboot):
            break
        v = struct.unpack_from("<I", eboot, off)[0]
        if i > 0 and (v < ptrs[-1] or v > dat_size * 1.05):
            break
        ptrs.append(v)
        i += 1
        if i > 40000:
            break
    return ptrs


def decompress_sce(container):
    """이벤트 컨테이너의 THEIRSCE(comptolib SCE) 블록을 해제해 표준 THEIRSCE 반환.
    JP는 타입1/3 압축. CN은 타입0(무압축)이라 그대로 통과."""
    sys.path.insert(0, PYTHONLIB_PATH)
    from pythonlib.utils import comptolib
    t = container.find(b"THEIRSCE")
    if t < 0:
        return None
    for hdr in range(t - 9, t - 40, -1):
        if hdr < 0:
            break
        typ = container[hdr]
        if typ not in (0, 1, 3):
            continue
        csz, dsz = struct.unpack_from("<II", container, hdr + 1)
        if not (0 < dsz < 5_000_000 and 0 < csz < 2_000_000):
            continue
        try:
            dec = comptolib.decompress_data(container[hdr:hdr + 9 + csz])
            if dec[:8] == b"THEIRSCE":
                return dec
        except Exception:
            pass
    # 타입0 무압축(헤더 없이 THEIRSCE부터 시작하는 케이스) 대비
    return container[t:]


def classify(speaker, voice):
    if voice:
        return "대사(음성)"
    if speaker:
        return "대사"
    return "표시/시스템"


def extract_scene(mini, Theirsce, container):
    """컨테이너 -> THEIRSCE 해제 -> Strings 섹션 엔트리 목록.
    반환: [{id, speaker, voice, jp}] (make_template과 동일 기준: Strings, Id!=-1)."""
    # 중요: mini.id는 씬 간 리셋되지 않고 누적됨(get_xml_from_theirsce 버그).
    # 씬마다 1로 리셋해야 fresh 추출(make_template/기존 JSON/map_jp_reference)과 Id 일치.
    mini.id = 1
    block = decompress_sce(container)
    if block is None or block[:8] != b"THEIRSCE":
        return None
    from lxml import etree
    theirsce = Theirsce(block)
    xml_bytes = mini.get_xml_from_theirsce(theirsce, "Story")
    root = etree.fromstring(xml_bytes)
    sroot = root.find(".//Strings")
    entries = (sroot.findall(".//Entry") if sroot is not None
               else root.findall(".//Entry"))
    rows = []
    for e in entries:
        ide = e.findtext("Id")
        if ide is None or ide.strip() == "-1":
            continue
        rows.append({
            "id": ide.strip(),
            "speaker": (e.findtext("SpeakerId") or "").strip(),
            "voice": (e.findtext("VoiceId") or "").strip(),
            "jp": (e.findtext("JapaneseText") or ""),
        })
    return rows


def load_existing_kr(trans_dir, scene):
    p = Path(trans_dir) / f"{scene}.json"
    if not p.exists():
        return {}
    data = json.load(open(p, encoding="utf-8"))
    return {str(l["id"]): (l.get("kr") or "") for l in data.get("lines", [])}


def cmd_export(args):
    sys.path.insert(0, PYTHONLIB_PATH)
    from story_pipeline_bin import make_mini
    from pythonlib.formats.rebirth.theirsce import Theirsce
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    mini = make_mini(args.tbl)

    dat = Path(args.dat).read_bytes()
    ptrs = read_pointers(args.eboot, len(dat))

    if args.all:
        scenes = [i for i in range(len(ptrs) - 1)
                  if b"THEIRSCE" in dat[ptrs[i]:ptrs[i + 1]]]
    else:
        scenes = [int(s) for s in args.scenes.split(",") if s.strip()]
    print(f"[i] 대상 씬 {len(scenes)}개")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "THEIRSCE"
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    scene_fill = PatternFill("solid", fgColor="D9E1F2")
    bold_w = Font(bold=True, color="FFFFFF")
    bold_b = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        ws.cell(1, c).font = bold_w
        ws.cell(1, c).fill = hdr_fill
    ws.freeze_panes = "A2"

    total_lines = 0
    empty_scenes = 0
    for n, sc in enumerate(scenes):
        try:
            rows = extract_scene(mini, Theirsce, dat[ptrs[sc]:ptrs[sc + 1]])
        except Exception as e:
            print(f"  [!] 씬 {sc} 추출실패: {type(e).__name__}: {e}", file=sys.stderr)
            continue
        if not rows:
            empty_scenes += 1
            continue
        kr_map = load_existing_kr(args.trans_dir, sc)
        # 씬 헤더행: 파일명(PS2 Story번호) + 씬번호 + 첫 문자열을 제목처럼
        title = next((r["jp"].replace("\n", " ") for r in rows if r["jp"].strip()
                      and not r["jp"].startswith("<")), "")
        hr = ws.max_row + 1
        ws.append([sc, "", "", "", f"{sc + STORY_OFFSET}.xml", title, ""])
        for c in range(1, len(COLS) + 1):
            ws.cell(hr, c).font = bold_b
            ws.cell(hr, c).fill = scene_fill
        for r in rows:
            ws.append([sc, r["id"], r["speaker"], r["voice"],
                       classify(r["speaker"], r["voice"]),
                       r["jp"], kr_map.get(r["id"], "")])
            ws.cell(ws.max_row, 6).alignment = wrap
            ws.cell(ws.max_row, 7).alignment = wrap
            total_lines += 1
        if (n + 1) % 50 == 0:
            print(f"    ...{n+1}/{len(scenes)} 씬 처리")

    widths = [8, 6, 10, 10, 12, 55, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(args.out)
    print(f"[OK] 저장: {args.out}  (대사 {total_lines}줄, 빈 씬 {empty_scenes}개 제외)")


def cmd_import(args):
    import openpyxl
    wb = openpyxl.load_workbook(args.inp, read_only=True)
    ws = wb.active
    # 헤더행에서 열 인덱스
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {name: header.index(name) for name in COLS if name in header}
    scenes = {}  # scene -> {id: line}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sc = row[idx["Scene"]] if idx.get("Scene") is not None else None
        idv = row[idx["Id"]] if idx.get("Id") is not None else None
        if sc is None or idv is None or str(idv).strip() == "":
            continue  # 헤더/빈행
        try:
            sc = int(sc)
        except (ValueError, TypeError):
            continue
        line = {
            "id": str(idv).strip(),
            "speaker": str(row[idx["Speaker"]] or "").strip() if "Speaker" in idx else "",
            "voice": str(row[idx["Voice"]] or "").strip() if "Voice" in idx else "",
            "jp": row[idx["Japanese"]] or "" if "Japanese" in idx else "",
            "kr": (row[idx["Korean"]] or "") if "Korean" in idx else "",
        }
        scenes.setdefault(sc, {})[line["id"]] = line

    Path(args.trans_dir).mkdir(parents=True, exist_ok=True)
    written = kr_total = 0
    for sc, lines_by_id in sorted(scenes.items()):
        p = Path(args.trans_dir) / f"{sc}.json"
        # 기존 순서 보존: 있으면 병합, 없으면 엑셀 순서.
        # 엑셀에만 있고 기존 JSON에 없는 id(예: 대사 JSON에 없던 시스템 문자열)는
        # 뒤에 추가한다. inject는 id로 찾으므로 순서는 빌드에 무관.
        if p.exists():
            data = json.load(open(p, encoding="utf-8"))
            have = {str(l["id"]) for l in data["lines"]}
            for l in data["lines"]:
                k = str(l["id"])
                if k in lines_by_id:
                    l["kr"] = lines_by_id[k]["kr"]
            for k, l in lines_by_id.items():
                if k not in have:
                    data["lines"].append(l)
        else:
            data = {"scene": sc, "lines": list(lines_by_id.values())}
        json.dump(data, open(p, "w", encoding="utf-8"),
                  ensure_ascii=False, indent=1)
        written += 1
        kr_total += sum(1 for l in data["lines"] if (l.get("kr") or "").strip())
    print(f"[OK] {written}개 씬 JSON 갱신, kr 채워진 줄 총 {kr_total}")


def cmd_split(args):
    """기존 엑셀에서 지정 Type 행만 골라 새 엑셀 생성. 씬 헤더행/구조 유지.
    Scene/Id/열구조 그대로라 import 왕복 그대로 됨."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    keep = set(t.strip() for t in args.types.split(","))
    src = openpyxl.load_workbook(args.inp, read_only=True).active
    rows = list(src.iter_rows(values_only=True))
    header = list(rows[0])
    ti = header.index("Type") if "Type" in header else 4
    ii = header.index("Id") if "Id" in header else 1

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "대사"
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    scene_fill = PatternFill("solid", fgColor="D9E1F2")
    bw = Font(bold=True, color="FFFFFF"); bb = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(1, c).font = bw; ws.cell(1, c).fill = hdr_fill
    ws.freeze_panes = "A2"

    kept = 0
    # 씬 헤더행은 Id가 비어있음 -> 그 씬에 남는 대사가 있을 때만 출력(지연 append)
    pending_hdr = None
    for r in rows[1:]:
        idv = r[ii]
        is_header = (idv is None or str(idv).strip() == "")
        if is_header:
            pending_hdr = r
            continue
        if r[ti] not in keep:
            continue
        if pending_hdr is not None:
            hr = ws.max_row + 1
            ws.append(list(pending_hdr))
            for c in range(1, len(header) + 1):
                ws.cell(hr, c).font = bb; ws.cell(hr, c).fill = scene_fill
            pending_hdr = None
        ws.append(list(r))
        ws.cell(ws.max_row, 6).alignment = wrap
        ws.cell(ws.max_row, 7).alignment = wrap
        kept += 1

    widths = [8, 6, 10, 10, 12, 55, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(args.out)
    print(f"[OK] {args.out}: Type {keep} 만 {kept}줄 추출")


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    e = sub.add_parser("export", help="DAT -> 엑셀")
    e.add_argument("--out", required=True)
    e.add_argument("--all", action="store_true", help="THEIRSCE 있는 모든 씬")
    e.add_argument("--scenes", default="", help="쉼표구분 씬번호")
    e.add_argument("--dat", default="DAT.BIN")
    e.add_argument("--eboot", default="ULJS00132_EBOOT.BIN")
    e.add_argument("--tbl", default="tbl_all.json")
    e.add_argument("--trans-dir", default="translation")
    e.set_defaults(func=cmd_export)

    im = sub.add_parser("import", help="엑셀 -> 번역JSON")
    im.add_argument("--in", dest="inp", required=True)
    im.add_argument("--trans-dir", default="translation")
    im.set_defaults(func=cmd_import)

    sp = sub.add_parser("split", help="엑셀에서 특정 Type만 골라 별도 엑셀 생성")
    sp.add_argument("--in", dest="inp", required=True)
    sp.add_argument("--out", required=True)
    sp.add_argument("--types", default="대사,대사(음성)",
                    help="쉼표구분 Type (기본: 대사,대사(음성))")
    sp.set_defaults(func=cmd_split)

    args = ap.parse_args()
    if args.cmd == "export" and not args.all and not args.scenes:
        print("[!] --all 또는 --scenes 필요", file=sys.stderr)
        sys.exit(1)
    args.func(args)


if __name__ == "__main__":
    main()
