# -*- coding: utf-8 -*-
# 미참조(시스템) 문자열 복원 재빌드: orddup + 전 미참조 서열 삽입(번역 치환).
#  텍스트 런만 번역으로 바꾸고 제어/태그 바이트는 JP 원본 그대로 보존.
#  사용: py sysrestore.py <scene...> [--dry]
import os
import sys
import json
import glob
import re
import struct

os.chdir(r"D:\clean_project")
sys.path.insert(0, r"D:\PythonLib"); sys.path.insert(0, ".")
from pathlib import Path
from lxml import etree
from pythonlib.formats.rebirth.scpk import Scpk
from pythonlib.formats.rebirth.theirsce import Theirsce
from pythonlib.utils import comptolib
from story_pipeline_bin import make_mini
from build_scene import inject_translation
from build_all_jp import parse_blobs
import importlib.util as ilu

DRY = "--dry" in sys.argv
# 고정오프셋 참조 블록(램 덤프 실증): 해당 rel 범위는 JP 오프셋에 pin
SCENE_FORCE = {  # 'ALL' = 미참조 풀 전체 pin (여관 앵커 없는 시설 전용 씬용)
    # 'PINALL' = ★참조 문자열까지 전부 JP 원위치 pin (2026-08-02, 씬4762 문 조사 메시지:
    #   참조 엔트리인데도 엔진이 고정 오프셋으로 읽어 '이대로는'의 첫 글자가 잘렸음.
    #   kr <= JP 슬롯 필수 - 초과 시 problems 보고)
    4198: "ALL", 5155: "ALL", 5156: "ALL", 5157: "ALL",
    5158: "ALL", 5159: "ALL", 5160: "ALL", 5161: "ALL", 5162: "ALL",
    4762: ("PINALL", [(1, 808)]),   # 문 조사 메시지(rel60/332/422/675) 원위치 고정
    4661: [(1101, 1488)],           # 구슬 풍선 힌트 5합성문 = 고정 오프셋 판독 (pin 필수)
    # 씬4725 질바 이름판: 화자 지정 없는 창이 이름을 고정 오프셋으로 읽어 공란
    #  (KR 풀에서 3B 앞으로 밀렸음) -> 이름 문자열만 JP 원위치 pin
    4725: ("PINALL", [(360, 375), (600, 615), (1620, 1635), (2080, 2095), (2725, 2740)]),
    # 씬5167 라질다항 음식 나눔 select: 식재 조각(플레인+Red쌍)을 식재ID별 고정 오프셋으로
    #  읽는 동적 select - 조각 미번역+오프셋 밀림으로 select 깨짐/소프트락. 전 블록 pin.
    #  ★イカ(4B)만 오징어(6B) 슬롯 초과라 JP 유지(가나 = 정상 렌더).
    5167: ("PINALL", [(368, 1240)]),
    # 가나 자유입력 퀴즈 5씬 (2026-08-04): 비교어 리스트+프롬프트+입력에코가 고정 오프셋
    #  판독인데 KR 풀에서 -86~-1363B 밀려 있었음 (4592 프리필 '거야!' = JP rel6040 위치의
    #  밀린 한글 꼬리 실증). 프롬프트~비교어~에코 블록 전체 JP 원위치 pin.
    4252: ("PINALL", [(2176, 2730)]),
    4448: ("PINALL", [(4194, 4268)]),
    # 4563/4585/4592 는 pin 미적용: 비교어/에코/프롬프트는 전부 XML 추적 포인터로 갱신됨을
    #  실측(퀴즈 판정은 pin 없이 정상 - 리프트 퀴즈 통과 실증). 문제는 '기본입력' F8 오퍼랜드가
    #  JP 널(프롬프트 직전)을 미추적 고정 참조 -> KR 풀에서 한글 꼬리('거야!' 프리필)를 읽던 것.
    #  해법 = STALE_OPS (아래): pin 강제 널런의 압축비용(+63B 등) 없이 오퍼랜드만 갱신.
    #  4252/4448 은 pin 이 슬롯에 들어가므로 유지(스테일 참조까지 JP 레이아웃으로 보장).
}
# ★'기본입력' 프리필: {씬: [(코드위치, 정답문자열 orig rel)]} -> u16 = newoff(정답)
#  입력창이 그 문자열로 미리 채워진 채 열림(플레이어는 결정만) - 가나 철자를 알 수 없는
#  한국어판 구제책. STALE_OPS(널=빈칸)와 같은 오퍼랜드, 값만 다름.
#  ★4563 실측 = 실패: 0x31fa/0x3204 (둘 다 프롬프트 직전 널 참조) 어느 쪽을 정답에 물려도
#   입력창은 공란 그대로. 즉 이 오퍼랜드는 '기본입력' 소스가 아님(4592 '거야!' 사례와 별개 경로).
#   -> 프리필 방식 보류, 프롬프트 병기로 대체. 재시도 시 런타임 추적(입력 버퍼 write BP) 필요.
#  ★3차 실패 원인 = 오프바이원: 인자 u16 은 0x3213 인데 0x3212(= f8 오퍼코드 자리)에 기록해
#   오퍼코드를 파괴 -> 입력칸 깨짐. 가설(마지막 f8 인자 = 기본입력) 자체는 미검증이었음.
#  4563 syscall 인자 구조: c0 | f8 [u16@0x3210 = 프롬프트 9927] | f8 [u16@0x3213 = 9954] | c8 04
PREFILL = {
    #  구조: 입력 syscall `c8 NN c1 e0 88 80` 직전 f8 인자 중 '프롬프트({0B}xx 합성문)'가 아닌
    #  널/템플릿 ref = 기본입력. 4563 실기검증 통과(2026-08-06).
    4563: [(0x3213, 9982)],    # タロット
    4252: [(0x1e3d, 2286)],    # フォルス (Q1. Q2 こおり는 syscall 공유라 프리필 불가 - id140 병기)
    4448: [(0x11fc, 4242)],    # いとしきそら
    4585: [(0x5de7, 16492)],   # はなのさくきせつ (JP 원판도 はなのさく 프리필 설계)
    4592: [(0x14cd, 6130)],    # もんしょう
}
# 미추적 '기본입력' 오퍼랜드: {씬: [(코드위치, 앵커 그룹 orig rel)]} -> u16 = newoff(앵커)-1 (널)
STALE_OPS = {
    4592: [(0x14cd, 6040)],
    4563: [(0x31fa, 9927), (0x3204, 9927)],
}
SCENES = [int(a) for a in sys.argv[1:] if a.isdigit()]

spec = ilu.spec_from_file_location("sp", "work/synopsis_jp/_spaced_inplace.py")
mod = ilu.module_from_spec(spec); spec.loader.exec_module(mod)
INIT = open("work/synopsis_jp/lzss_init.bin", "rb").read()
TKR = {k.lower(): v for k, v in json.load(open("tbl_full_kr.json", encoding="utf-8"))["TBL"].items()}
INVK = {v: k for k, v in TKR.items()}
TALL = {k.lower(): v for k, v in json.load(open("tbl_all.json", encoding="utf-8"))["TBL"].items()}
NM = json.load(open("work/names_npc.json", encoding="utf-8"))
TAG_RE = re.compile(r"<[^>]*>")


def is_translatable(run):
    # 가나/한자/한글 포함 런만 번역 대상 (숫자/기호/@/말줄임 단독 등은 제외)
    for ch in run:
        o = ord(ch)
        if 0x3040 <= o <= 0x30FF or 0x4E00 <= o <= 0x9FFF or 0xAC00 <= o <= 0xD7AF:
            return True
    return False


def norm_text(s):
    # 시트/JSON 텍스트 -> 번역대상 텍스트 런 리스트 (태그 제거, 개행 분리)
    s = TAG_RE.sub("\n", s)
    runs = [r for r in (x.strip().strip(";") for x in s.split("\n")) if r]
    return runs


def key_runs(runs):
    return tuple(r for r in runs if is_translatable(r))


# ---- 전역 번역 사전: (jp 텍스트런 튜플) -> kr 텍스트런 리스트 ----
SYSDICT = {}
JOINDICT = {}


def learn(jp, kr):
    if not jp or not kr:
        return
    jr = key_runs(norm_text(jp)); krr = key_runs(norm_text(kr))
    if jr and krr:
        SYSDICT.setdefault(jr, list(krr))


import openpyxl
wb = openpyxl.load_workbook("tor_system.xlsx", read_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True); next(it)
for r in it:
    learn(str(r[5] or ""), str(r[6] or ""))
wb.close()
for f in glob.glob("translation/*.json"):
    d = json.load(open(f, encoding="utf-8"))
    for ln in d["lines"]:
        learn(ln.get("jp") or "", ln.get("kr") or "")
for jp_name, kr_name in NM.items():
    learn(jp_name, kr_name)
for _k,_v in SYSDICT.items():
    JOINDICT.setdefault("".join(_k).replace("　", ""), _v)
print(f"[사전] {len(SYSDICT)}개 항목 / 조인 {len(JOINDICT)}")


def tokenize(b):
    # JP 바이트 -> [('ctrl',bytes) | ('text',str)] 토큰. 텍스트 = TALL 2바이트 or ASCII 인쇄문자.
    toks = []; i = 0
    while i < len(b):
        c = b[i]
        if 0x20 <= c < 0x7F:
            ch = chr(c)
            if toks and toks[-1][0] == "text":
                toks[-1] = ("text", toks[-1][1] + ch)
            else:
                toks.append(("text", ch))
            i += 1
            continue
        h2 = b[i:i + 2].hex()
        if c >= 0x80 and h2 in TALL:
            ch = TALL[h2]
            if toks and toks[-1][0] == "text":
                toks[-1] = ("text", toks[-1][1] + ch)
            else:
                toks.append(("text", ch))
            i += 2
            continue
        if toks and toks[-1][0] == "ctrl":
            toks[-1] = ("ctrl", toks[-1][1] + bytes([c]))
        else:
            toks.append(("ctrl", bytes([c])))
        i += 1
    return toks


INVA = {}
for _k, _v in TALL.items():
    INVA.setdefault(_v, _k)


def enc_kr_text(s):
    o = bytearray()
    for c in s:
        if c == " ":
            o.append(0x20)
        elif ord(c) < 0x7F:
            o.append(ord(c))
        elif c in INVK:
            o.extend(bytes.fromhex(INVK[c]))
        elif c in INVA:
            # ★색 파라미터 잔재 등 kr 에 남겨야 하는 JP 문자(金$ 류)는 tbl_all 로 통과
            #  (2026-08-02, 4661 구슬 힌트 색코드 실증)
            o.extend(bytes.fromhex(INVA[c]))
        else:
            return None
    return bytes(o)


# 반말/타운별 여관 주인장 변형 = 이 씬(미나르, 존댓말 여주인)에선 미사용 - JP 유지(압축 절약)
SKIP_SUBSTR = ("泊まんのかい", "ガルド置いてきな", "だけど泊まるかい", "だけど、　泊まるかい",
               "ガルドだ。", "泊まらないのか", "ガルドだけど、", "またおいで", "ごゆっくり。",
               "お金が足りないよ。", "お金が足りないぞ。", "行ってラッシャーイ", "泊まる？",
               "お金が足りないみたいね", "お金が足りないんじゃが", "いつでもおいで",
               "よく眠れたかい", "よく眠れた？", "いつでもどうぞ", "ありがとうね",
               "いってらっしゃい。", "またな……")

MANUAL = {  # (jp 번역런 튜플) -> kr 런 리스트 (조각형/색체인 수동 등록)
    # 매지컬포트 레벨업 확인창 둘째 줄 (슬롯 26B 적합 축약, 2026-08-01)
    ("間隔で食材が配達されます。",): ["간격으로 배달됩니다."],
    ("種類の食材が配達されます。",): ["종류의 식재가 배달됩니다."],
    ("食材の個数が増加します。",): ["식재 개수가 증가합니다."],
    ("ヒト",): ["사람"],
    ("お金が足りないようです。",): ["돈이 부족한 것 같습니다."],
    ("道中、お気をつけて。",): ["조심해서 가세요."],
    ("よ。", "怖いわけないわ。", "真剣に話せば気持ちは通じるはずよ。"):
        ["이야.", "무서울 리 없어.", "진심으로 말하면 마음은 통할 거야."],
    ("よ。", "怖がってはいけないわ。"): ["이야.", "무서워하면 안 돼."],
    ("スズの樹",): ["스즈나무"],
    ("だ。", "この実を使って作る石けんは、", "いい香りがしてお肌にもいいんだよなぁ。"):
        ["야.", "이 열매로 만든 비누는,", "향도 좋고 피부에도 좋대."],
    ("に聞いてみよっと。",): ["한테 물어봐야지."],
    ("『ゴルドバの月夜』", "百年前の"): ["『골드바의 달밤』", "백 년 전"],
    ("ゴルドバの日",): ["골드바의 날"],
    ("の", "夜の海岸を描いた歴史的作品。", "プロポーズにふさわしい月夜だった。"):
        ["의", "밤 해안을 그린 명작.", "청혼에 어울리는 달밤이었다."],
    ("えっ!?", "クレアさん……？"): ["엣!?", "클레어 씨…?"],
    ("……待って、",): ["…잠깐만,"],
    ("あなた、手にケガしてるでしょ？",): ["너, 손 다쳤지?"],
    ("食材を買う", "世間話をする", "やめる"):
        ["식재료를 산다", "이야기를 나눈다", "그만둔다"],
    ("泊まる", "世間話をする", "やめる"):
        ["묵는다", "이야기를 나눈다", "그만둔다"],
    # 이벤트 리플레이 카테고리 라벨 (2026-08-02 전수복원 때 미사전이라 JP 잔존하던 것)
    ("二回目以降の発見",): ["두 번째 이후의 발견"],
    # 씬5167 음식 나눔 식재 조각 (식재 아이템명 표준 준수. イカ 는 슬롯초과라 미등록=JP 유지)
    ("ライス",): ["라이스"],
    ("ポーク",): ["포크"],
    ("ジャガイモ",): ["감자"],
    ("ニンジン",): ["당근"],
    ("カレーの材料",): ["카레 재료"],
    ("タコ",): ["문어"],
    ("マグロ",): ["참치"],
    ("タイ",): ["도미"],
    ("シャケ",): ["연어"],
    ("エビ",): ["새우"],
    ("シーフード",): ["해물"],
    # 씬4661 발카 수용소 구슬 퍼즐 힌트 (태그인식 워크가 <06> u32 파라미터로 널을 삼켜
    #  블록이 '합성문 단위'로 잡힘 - 키는 합성 런 튜플. '3 '/'``` '/'金$ ' = 색 파라미터
    #  잔재로 원바이트 보존 필수 - enc_kr_text 의 INVA 폴백이 처리)
    #  ★pin 슬롯 예산상 축약 표현(메모와 동일 문형): 왼쪽은/왼쪽엔, 꼬리 」
    #  색이름 런의 '3 /``` /金$ ' 뒤 공백 1개 추가 = 조사(은/엔)와 색이름 사이 실공백
    #  (JP 는 に+黄 붙는 구조라 직역 시 '왼쪽엔노랑' - 1차 실기서 확인)
    ("ギンナルのヒント", "赤い球", "左隣は", "3 黄", "だ」"):
        ["긴나르 힌트", "빨간 구슬", "왼쪽은", "3  노랑", "」"],
    ("ギンナルのヒント", "``` 灰色の球", "左隣は", "金$ 茶", "だ」"):
        ["긴나르 힌트", "``` 회색 구슬", "왼쪽은", "金$  갈색", "」"],
    ("ギンナルのヒント", "3 黄色い球", "左方向に", "``` 灰", "はない」"):
        ["긴나르 힌트", "3 노란 구슬", "왼쪽엔", "```  회색", "은 없다」"],
    ("ギンナルのヒント", "青い球", "左方向に", "3 黄", "はない」"):
        ["긴나르 힌트", "파란 구슬", "왼쪽엔", "3  노랑", "은 없다」"],
    ("ギンナルのヒント", "金$ 茶色い球", "カスれて読めない……"):
        ["긴나르 힌트", "金$ 갈색 구슬", "지워져 못 읽는다……"],
}


def translate_unref(b):
    """미참조 JP 문자열 바이트 -> 번역 치환 바이트 (실패시 None = JP 원본 유지).
    번역대상 런(가나/한자)만 kr 치환, 제어/태그/숫자/기호 런은 원본 바이트 그대로."""
    toks = []
    i = 0
    while i < len(b):
        c = b[i]
        if 0x20 <= c < 0x7F:
            st = i
            while i < len(b) and 0x20 <= b[i] < 0x7F:
                i += 1
            toks.append(["text", bytes(b[st:i])])
            continue
        if c >= 0x80 and b[i:i + 2].hex() in TALL:
            st = i
            while i + 1 < len(b) and b[i] >= 0x80 and b[i:i + 2].hex() in TALL:
                i += 2
            toks.append(["text", bytes(b[st:i])])
            continue
        st = i
        while i < len(b):
            c2 = b[i]
            if 0x20 <= c2 < 0x7F:
                break
            if c2 >= 0x80 and b[i:i + 2].hex() in TALL:
                break
            i += 1
        toks.append(["ctrl", bytes(b[st:i])])
    merged = []
    for t in toks:
        if merged and merged[-1][0] == t[0]:
            merged[-1][1] += t[1]
        else:
            merged.append(t)
    toks = merged

    def dectxt(bb):
        s = ""; j = 0
        while j < len(bb):
            if bb[j] < 0x80:
                s += chr(bb[j]); j += 1
            else:
                s += TALL[bb[j:j + 2].hex()]; j += 2
        return s

    tr_idx = [k for k, (kind, val) in enumerate(toks)
              if kind == "text" and is_translatable(dectxt(val))]
    if not tr_idx:
        return None
    key = tuple(dectxt(toks[k][1]).strip(" ;@　") for k in tr_idx)

    kr = MANUAL.get(key) or SYSDICT.get(key)
    if kr is None:
        kr = JOINDICT.get("".join(key).replace("　", ""))
    if kr is None:
        return None
    if len(kr) != len(tr_idx):
        if len(tr_idx) == 1:
            kr = [" ".join(kr)]
        else:
            return None
    out = bytearray(); ki = 0
    for k, (kind, val) in enumerate(toks):
        if kind == "ctrl":
            out.extend(val)
            continue
        if k not in tr_idx:
            out.extend(val)
            continue
        txt = dectxt(val)
        lead = txt[:len(txt) - len(txt.lstrip(" ;@　"))]
        tn = len(txt) - len(txt.rstrip(" @　"))
        trail = txt[len(txt) - tn:] if tn else ""
        eb = enc_kr_text(kr[ki])
        if eb is None:
            return None
        out.extend((enc_kr_text(lead) or b"") + eb + (enc_kr_text(trail) or b""))
        ki += 1
    return bytes(out)


def rp(buf, ds):
    p = []; k = 0
    while True:
        v = struct.unpack_from("<I", buf, 0x126F90 + k * 4)[0]
        if k > 0 and (v < p[-1] or v > ds * 1.05):
            break
        p.append(v); k += 1
        if k > 40000:
            break
    return p


src = open("DAT.BIN", "rb").read()
sp_ = rp(open("ULJS00132_EBOOT.BIN", "rb").read(), len(src))
dat = bytearray(open("DAT_jp_final.BIN", "rb").read())
dp = rp(open("EBOOT_jp_new.BIN", "rb").read(), len(dat))
mini_ex = make_mini("tbl_all.json")
mini_in = make_mini("tbl_full_kr.json")

for SC in SCENES:
    p0 = sp_[SC]; base = src.rfind(b"SCPK", max(0, p0 - 64), p0 + 8)
    nf = struct.unpack_from("<I", src, base + 8)[0]
    sizes = [struct.unpack_from("<I", src, base + 16 + 4 * k)[0] for k in range(nf)]
    scont = bytes(src[base:base + 16 + 4 * nf + sum(sizes)])
    Path("work/_nb.bin").write_bytes(scont)
    scpk = Scpk.from_path(Path("work/_nb.bin")); orig = scpk.rsce
    so = Theirsce(orig).strings_offset
    mini_ex.id = 1
    Path("work/_nb.xml").write_bytes(mini_ex.get_xml_from_theirsce(Theirsce(orig), "Story"))
    d = json.load(open(f"translation/{SC}.json", encoding="utf-8"))
    inject_translation("work/_nb.xml", "work/_nbk.xml", d.get("lines", []))
    tree = etree.parse("work/_nbk.xml"); root = tree.getroot()
    for e in root.findall(".//Speakers/Entry"):
        jt = e.find("JapaneseText"); et = e.find("EnglishText")
        if jt is not None and et is not None and (jt.text or "") in NM:
            et.text = NM[jt.text]
    mini_in.id = 1
    ents = [e for e in root.iter("Entry")
            if e.find("Id") is not None and e.find("Id").text != "-1"
            and e.find("PointerOffset") is not None
            and e.find("PointerOffset").text not in (None, "-1")]
    groups = {}; normset = set()
    for e in ents:
        for x in e.find("PointerOffset").text.split(","):
            p = int(x)
            o = struct.unpack_from("<H", orig, p)[0]
            groups.setdefault(o, [e, []])[1].append(p)
            if o == 0:
                continue   # ptr=0 관용(빈 방출) - normset 등록 시 풀 머리 실문자열이 탈락(4622 실증)
            o2 = o
            while so + o2 < len(orig) and orig[so + o2] == 0:
                o2 += 1
            # ★빈 엔트리가 '앞널 경유'로 실문자열을 가리키면 normset 등록 금지 (2026-08-02):
            #  방출이 빈 문자열이라 실문자열이 통째 탈락(5029 rel109/369/19144 실증)
            #  -> unref 복원(JP 원바이트)에 맡김.
            jt_ = e.find("JapaneseText")
            if o2 != o and (jt_ is None or not jt_.text):
                continue
            normset.add(o2)
    # 미포착 이름 오퍼랜드 (일반해)
    known = {p for _, (e_, ptrs_) in groups.items() for p in ptrs_}
    q = 0
    while True:
        q = orig.find(b"\x48\x20\x04\xf8", q, so - 6)
        if q < 0:
            break
        p = q + 4
        v = struct.unpack_from("<H", orig, p)[0]
        tgt_ok = so + v < len(orig) and orig[so + v] != 0 and (v == 0 or orig[so + v - 1] == 0)
        if p not in known and tgt_ok:
            if v in groups:
                groups[v][1].append(p)
            else:
                st = so + v; n = st
                while n < len(orig) and orig[n] != 0:
                    n += 1
                kb = translate_unref(orig[st:n]) or bytes(orig[st:n])
                groups[v] = [("RAW", kb), [p]]
        q += 1
    # 전 미참조 = 서열 삽입 (번역 치환)
    # ★태그인식 문자열 경계 (2026-08-01): <05>/<06>/<07>/<0E> 는 u32 파라미터를 소비하므로
    #  파라미터 안의 널을 문자열 경계로 오인하면 안 됨(추출기 bytes_to_text 와 동일 파서 사용).
    #  종전 널단위 분해는 select 창 합성문 조각을 별개 문자열로 중복 복원 -> 서수 밀림(실증).
    tio = Theirsce(orig)
    strs = []
    q = so
    while q < len(orig):
        if orig[q] == 0:
            q += 1; continue
        st_abs = q
        try:
            mini_ex.bytes_to_text(tio, st_abs)
            en_abs = tio.tell() - 1
        except Exception:
            en_abs = st_abs
        if en_abs <= st_abs or en_abs > len(orig):
            en_abs = st_abs
            while en_abs < len(orig) and orig[en_abs] != 0:
                en_abs += 1
        strs.append((st_abs - so, en_abs))
        q = en_abs + 1
    starts = [s for s, _ in strs]
    ends = {s: e for s, e in strs}   # rel_start -> abs_end(널 직전)
    # ★문자열 중간 참조 그룹(씬판 접미사 포인터, 5155 食材調達施設+4=調達施設 실증):
    #  standalone 사본 방출 금지(서수 +1 = select 밀림) -> 호스트 내부로 리다이렉트.
    #  kr 호스트는 JP 와 같은 바이트 오프셋에 접미사가 오도록 번역 정렬 필요.
    midfix = {}
    for o in [k for k in groups.keys() if k > 0 and orig[so + k - 1] != 0]:
        h = max((s for s in starts if s <= o), default=None)
        if h is not None and h in ends and so + o < ends[h]:
            midfix[o] = (h, o - h, groups[o][1])
            del groups[o]
            print(f"  [중간참조] rel{o} -> 호스트 rel{h}+{o - h}")
    unref = [s for s in starts if s not in groups and s not in normset]
    stats = {"tr": 0, "raw": 0}
    FORCE = SCENE_FORCE.get(SC)
    PINALL = False
    if isinstance(FORCE, tuple) and FORCE[0] == "PINALL":
        PINALL = True
        FORCE = FORCE[1]
        print(f"  [FORCE=PINALL] {FORCE}")
    elif FORCE == "PINALL":
        PINALL = True
        FORCE = [(1, len(orig) - so)]
        print(f"  [FORCE=PINALL] {FORCE}")
    elif FORCE == "ALL":
        FORCE = [(1, len(orig) - so)]
        print(f"  [FORCE=ALL] {FORCE}")
    elif FORCE is None:
        # 자동 탐지: 여관 시스템 블록 = 첫 미참조 <0e><02> 조각부터, select/notice/스캐폴드
        #  (<06>@<01>/<07>) 를 만나기 전까지의 미참조 연속 구간
        askmark = [bytes.fromhex("".join(
            {v: k for k, v in TALL.items()}[c] for c in w)) for w in ("泊", "ガルド")]
        us = sorted(unref)
        lo = hi = None
        for idx, s in enumerate(us):
            st2 = so + s; n2 = ends[s]
            bb2 = bytes(orig[st2:n2])
            # 태그인식 경계에선 여관 ask = <0E>태그로 시작 + 泊/ガルド 포함 한 문자열
            if lo is None and bb2[:1] == b"\x0e" and any(m in bb2 for m in askmark):
                lo = s; hi = n2 - so
                break
            if lo is None:
                if bb2 == b"" and idx + 1 < len(us):
                    st3 = so + us[idx + 1]; n3 = st3
                    while n3 < len(orig) and orig[n3] != 0:
                        n3 += 1
                    nxtb = bytes(orig[st3:n3])
                    if any(m in nxtb for m in askmark):
                        lo = s; hi = n2 - so
            else:
                if bb2 in (b"select", b"notice", b"@", b"") or bb2.startswith(b"@"):
                    break
                hi = n2 - so
        # 여관 앵커 이후 전체(식재점/메모/레벨업 창 등 시스템 꼬리 전부)를 pin - 2026-08-01
        #  (스캐폴드 직전 중단 방식은 식재점 매지컬포트 창 누락 = select 깨짐 실증)
        FORCE = [(lo, len(orig) - so)] if lo is not None else []
        print(f"  [자동 FORCE-확장] {FORCE}")

    def in_force(o):
        return any(lo <= o < hi for lo, hi in FORCE)

    # 색체인 꼬리 조각(참조 엔트리의 합성문에 이미 포함된 미참조)은 종전대로 드롭
    #  (조립기는 합성문에서 멈추므로 미사용 - 복원하면 내용 중복으로 풀만 커짐).
    #  단 FORCE 범위는 고정오프셋 참조 실증 구역이라 무조건 복원.
    # ★ptr=0 팬텀 그룹(방출 안 됨)의 엔트리는 corpus 제외 (2026-08-02):
    #  맵제목 등이 ptr=0 로 XML 에 잡히면 corpus 에 들어가 is_chain_tail 이 미참조 원본을
    #  '이미 포함됨'으로 오판 -> 복원 드롭 -> 서수 밀림 (4684 실증). 방출되는 엔트리만 코퍼스로.
    emitted_ids = {id(e_g) for o_g, (e_g, _p) in groups.items()
                   if o_g != 0 and not isinstance(e_g, tuple)}
    ref_jp_joined = []
    for e in ents:
        if id(e) not in emitted_ids:
            continue
        jt = e.find("JapaneseText")
        if jt is not None and jt.text:
            ref_jp_joined.append("".join(key_runs(norm_text(jt.text))).replace("　", ""))
    corpus = "\n".join(ref_jp_joined)

    def is_chain_tail(o):
        if in_force(o):
            return False
        st = so + o; n = ends[o]
        bb = orig[st:n]
        txt = ""
        j = 0
        while j < len(bb):
            if bb[j] < 0x80:
                txt += chr(bb[j]); j += 1
            elif bb[j:j + 2].hex() in TALL:
                txt += TALL[bb[j:j + 2].hex()]; j += 2
            else:
                txt += "\n"; j += 1
        runs = "".join(key_runs(norm_text(txt))).replace("　", "")
        return bool(runs) and runs in corpus

    # ★chain_tail 드롭 폐지 (2026-08-02): 태그인식 워크가 별개 문자열로 세는 것은 전부
    #  복원해야 JP 서수 일치 (참조문과 동일 텍스트인 '중복 사본'도 JP 풀에 실존 - 4363 rel836
    #  실증). 중복은 LZ 매칭으로 압축비용 거의 0. is_chain_tail 은 참고용으로만 보존.
    _ = is_chain_tail  # (미사용)

    items = sorted(set(list(groups.keys()) + unref))
    out = bytearray(orig[:so + 1]); newoff = {}; allnew = {}; problems = []
    for o in items:
        if o == 0 and o in groups:
            newoff[0] = 0
            continue
        st = so + o
        if o in ends:
            n = ends[o]
        else:
            n = st
            while n < len(orig) and orig[n] != 0:
                n += 1
        if o in unref:
            kb = translate_unref(orig[st:n])
            if kb is None:
                kb = bytes(orig[st:n]); stats["raw"] += 1
            else:
                stats["tr"] += 1
        else:
            e, _ = groups[o]
            if isinstance(e, tuple) and e[0] == "RAW":
                kb = e[1]
            else:
                try:
                    kb = mini_in.get_node_bytes(e)
                except Exception:
                    kb = b""
        if in_force(o) and (o in unref or PINALL):
            cur = len(out) - so
            if cur > o:
                problems.append(f"rel{o}: pin 선행초과 +{cur - o}")
            else:
                out += b"\x00" * (o - cur)
                nxt2 = [s for s in starts if s > o]
                nxt = nxt2[0] if nxt2 else len(orig) - so
                if len(kb) > nxt - o - 1:
                    problems.append(f"rel{o}: kr {len(kb) - (nxt - o - 1)}B 슬롯초과")
        if o in groups:
            newoff[o] = len(out) - so
        allnew[o] = len(out) - so
        _tr = os.environ.get("SYSR_TRACE")
        if _tr:
            _a, _b = (map(int, _tr.split(","))) if "," in _tr else (6900, 7400)
        if _tr and _a <= o <= _b:
            gap_next = ([s for s in starts if s > o] + [len(orig) - so])[0] - o
            print(f"    o={o} {'unref' if o in unref else 'REF'} jpgap={gap_next} kr={len(kb)} cur={len(out)-so}")
        out += kb + b"\x00"
    for pr in problems[:25]:
        print("  " + pr)
    if problems and not DRY:
        print(f"  [STOP] scene {SC}: pin 문제 {len(problems)}건")
        continue
    for o, (e, ptrs) in groups.items():
        for p in ptrs:
            struct.pack_into("<H", out, p, newoff[o])
    for o, (h, delta, ptrs) in midfix.items():
        base = newoff.get(h, h)   # REF=이동 위치 / unref pin=JP 원위치
        for p in ptrs:
            struct.pack_into("<H", out, p, base + delta)
    for cp, ans in PREFILL.get(SC, []):
        assert ans in newoff, f"PREFILL 정답 rel{ans} 미발견"
        struct.pack_into("<H", out, cp, newoff[ans])
        print(f"  [PREFILL] 0x{cp:x} -> rel{newoff[ans]} (정답 프리필)")
    _pf = {cp for cp, _ in PREFILL.get(SC, [])}
    for cp, anc in STALE_OPS.get(SC, []):
        if cp in _pf:
            continue
        assert anc in newoff, f"STALE_OPS 앵커 rel{anc} 미발견"
        tgt = newoff[anc] - 1
        assert out[so + tgt] == 0, f"STALE_OPS 타깃 rel{tgt} 비널"
        struct.pack_into("<H", out, cp, tgt)
        print(f"  [STALE_OP] 0x{cp:x} -> rel{tgt} (널, 앵커 {anc}->{newoff[anc]})")
    # ★일반 오퍼랜드 스테일 갱신 (2026-08-04, 4597 꽃집아이 이름판 '.' 사건):
    #  `48 XX 04 f8 [u16]` 오퍼랜드 중 XML 미포착(= 위 groups 패스가 안 건드린) 것을 전수 갱신.
    #  조건: JP u16 이 JP 문자열 시작 && out 의 현재 값이 JP 값 그대로(스테일) && 새 위치 확보.
    #  (기존 _rebuild_scenes 일반해는 XX=0x20 만 스캔 - XX=0x24/0x28 등 이름판 오퍼랜드가 그물 밖)
    q2 = 0
    while True:
        q2 = orig.find(b"\x04\xf8", q2 + 1)
        if q2 < 0 or q2 + 4 > so:
            break
        if orig[q2 - 2] != 0x48:
            continue
        jv = struct.unpack_from("<H", orig, q2 + 2)[0]
        if jv not in allnew:
            continue
        cur = struct.unpack_from("<H", out, q2 + 2)[0]
        if cur == jv and allnew[jv] != jv:
            struct.pack_into("<H", out, q2 + 2, allnew[jv])
            print(f"  [OP갱신] 0x{q2-2:x} rel{jv} -> rel{allnew[jv]}")
    new = bytes(out).replace(b"\x0e\x02\x00\x00\x80", b"\x0e\x02\x00\x00\x00")  # jeongak x80 -> font:2
    ok_sel = new.count(b"select") == orig.count(b"select")
    print(f"scene {SC}: select {new.count(b'select')}/{orig.count(b'select')} pool {len(new)}B (JP {len(orig)}B) "
          f"미참조 번역 {stats['tr']} / 원문유지 {stats['raw']}")
    if DRY or not ok_sel:
        if not ok_sel:
            print("  [STOP] select 불일치")
        continue
    blobs = parse_blobs(scont)
    _, off, size, idx = next(bb for bb in blobs if bb[0] == "sce")
    blob = comptolib.compress_data(new, version=scpk._rsce_comp_type); how = "greedy"
    if len(blob) > size:
        body = mod.lzss_encode_optimal(new, INIT)
        blob = struct.pack("<b", 1) + struct.pack("<L", len(body)) + struct.pack("<L", len(new)) + body
        assert comptolib.decompress_data(blob) == new
        how = "optimal"
    if len(blob) > size:
        # type3 폴백 (JP THEIRSCE 는 type1/3 혼재 = 디스패처 동일. pin 널런에 유리)
        b3 = comptolib.compress_data(new, version=3)
        if len(b3) < len(blob) and comptolib.decompress_data(b3) == new:
            blob = b3; how = "type3"
    if len(blob) > size:
        print(f"  [FAIL] blob over by {len(blob)-size}B ({how})")
        continue
    newc = scont[:off] + blob + b"#" * (size - len(blob)) + scont[off + size:]
    q0 = dp[SC]; b2 = dat.rfind(b"SCPK", max(0, q0 - 64), q0 + 8)
    assert bytes(dat[b2:b2 + 4]) == b"SCPK"
    assert b2 + len(newc) <= dp[SC + 1]
    dat[b2:b2 + len(newc)] = newc
    print(f"  [OK] scene {SC} (blob {len(blob)}/{size}B, {how})")

if not DRY:
    open("DAT_jp_final.BIN", "wb").write(bytes(dat))
    print("[SAVED]")
